# coding: UTF-8

import openpyxl
from openpyxl.styles import Font, colors, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
import configparser
import jpholiday
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
import calendar
import copy
from emoji import emojize


# 使わせてもらったコード
# https://qiita.com/TakayukiKiyohara/items/83469bbf9d3786333f48
# https://openpyxl.readthedocs.io/en/stable/tutorial.html


# メイン処理
def main():

    # 開始ログお茶目ver
    print('月次処理を開始するよ！' + emojize(":thumbs_up:"))


    # 設定ファイル取得
    iniFile = getIniFile()

    # Excelファイルを読み込み
    wb = openpyxl.load_workbook(iniFile.get('settings', 'IN'))


    # 表のタイトルを修正
    sheet = wb[iniFile.get('settings', 'TARGET')]

    # 各タイトル部分
    for i in ['A1','A21','A41','A61','A81']:

        # タイトル部分を入力
        sheet[i].value = createTitle()

        # アピール用
        sheet[i].comment = Comment('プログラムで自動作成しました。','Author')




    # 初期化
    initTable(sheet,wb[iniFile.get('settings', 'TEMPLATE')])

    input3rdSunday(sheet)

    # changeColor(sheet)

    # テンプレ部分を非表示にする
    sheet.row_dimensions.group(102, 119, hidden=True)

    # 保存
    wb.save(iniFile.get('settings', 'OUT') + createTitle() + '.xlsx')

    # 終了ログお茶目ver
    print('終了したよ！ファイルを見てみよう！' + emojize(":sparkling_heart:"))


# 設定ファイル取得
def getIniFile():
    iniFile = configparser.ConfigParser()
    iniFile.read('./config.ini', 'UTF-8')
    return iniFile

#タイトルを生成
def createTitle():

    # 再来月のやつとか必要かも
    return str(nextMonth().month) + '月シフト'

# 来月の1日を取得
def nextMonth():

    # とりあえず翌月固定
    monthNum = 1

    # 来月の1日を取得
    return datetime.today().replace(day = 1) + relativedelta(months = monthNum)

# 表を翌月で初期化
def initTable(sheet,templateSheet):

    writeSchedule(sheet)



# 表に日付を記入
def writeSchedule(sheet):

    day    = nextMonth().date()
    rowNum = 2

    # 曜日と列番号の辞書
    dic = {0 : 6 , 1 : 11 , 2 : 16 , 3 : 21 , 4 : 26 , 5 : 31 , 6 : 36}

    for i in range(calendar.monthrange(day.year, day.month)[1]):

        # 曜日に応じた列番号を取得
        colNum = dic[day.weekday()]

        # 日付の記入
        sheet.cell(row = rowNum, column = colNum).value = day

        # 第3SUNDAYの場合
        if (colNum == 36 and rowNum == 42):

            input3rdSunday(sheet)

        # 祝日の場合
        elif jpholiday.is_holiday(day):

            # 曜日を祝日にする
            sheet.cell(row = rowNum + 1, column = colNum).value = '祝'

            # 日付と曜日欄の文字色を赤に
            sheet.cell(row = rowNum    , column = colNum).font = Font(color=colors.RED)
            sheet.cell(row = rowNum + 1, column = colNum).font = Font(color=colors.RED)

            # TODO 一律、日曜出勤メンバーを出す
            RangeCopyCell(sheet, 36, 104, 36 + 4, 104 + 15, colNum - 36, -(104 - 2 - rowNum))





        # それ以外
        else:

            # シフトのコピー
            RangeCopyCell(sheet, colNum, 104, colNum + 4, 104 + 15, 0, -(104 - 2 - rowNum))

        # 日付を1日足す
        day += relativedelta(days = 1)

        # 月曜日の場合
        if day.weekday() == 0:
            # TODO 次の行へ移動する
            rowNum += 20


# 範囲をコピー
def RangeCopyCell( sheet, min_col, min_row, max_col, max_row, shift_col, shift_row ):
    #コピー元の結合されたセルの結合を解除していく。
    merged_cells = copy.deepcopy(sheet.merged_cells)
    for merged_cell in merged_cells :
        if merged_cell.min_row >= min_row \
            and merged_cell.min_col >= min_col \
            and merged_cell.max_row <= max_row \
            and merged_cell.max_col <= max_col :
                #結合を解除していく。
                sheet.unmerge_cells(merged_cell.coord)

    #全セルをコピー
    for col in range( min_col, max_col + 1):
        for row in range( min_row, max_row + 1):
            fmt = "{min_col}{min_row}"
            #コピー元のコードを作成。
            copySrcCoord = fmt.format(
                min_col = get_column_letter(col),
                min_row = row )

            #コピー先のコードを作成。
            copyDstCoord = fmt.format(
                min_col = get_column_letter(col + shift_col),
                min_row = row + shift_row )
            #コピー先に値をコピー。
            if type(sheet[copySrcCoord]) != MergedCell :
                sheet[copyDstCoord].value = sheet[copySrcCoord].value
                #書式があったら、書式もコピー。
                if sheet[copySrcCoord].has_style :
                    sheet[copyDstCoord]._style = sheet[copySrcCoord]._style

    #結合解除したセルを再結合していく。
    for merged_cell in merged_cells :
        if merged_cell.min_row >= min_row \
            and merged_cell.min_col >= min_col \
            and merged_cell.max_row <= max_row \
            and merged_cell.max_col <= max_col :
                #結合していく。
                sheet.merge_cells(merged_cell.coord)
                #コピー先のセルの結合範囲情報を作成する。
                newCellRange = copy.copy(merged_cell)
                #コピー先のセルの結合範囲情報をずらす。ここではshiftRow分ずらしている。
                newCellRange.shift(shift_col, shift_row)
                sheet.merge_cells(newCellRange.coord)



# 第3SUNDAYのシフト変更用
def input3rdSunday(sheet):

    # 第3SUNDAYは第3日曜日なので、セル[AJ42]
    # TODO 13直指定
    for i in range(13):

        # TODO 固定値バリバリ…
        RangeCopyCell(sheet, 6, 105, 6 + 4, 105, 30, - (105 - (45 + i)))


        sheet.merge_cells(start_row = 59, start_column = 36, end_row = 59, end_column = 36 + 4)
        sheet.cell(row = 59, column = 36).value = '第3SUNDAY大集合'

def changeColor(sheet):

    # 全セルを確認
    for row in sheet:
        for cell in row:
            if sheet[cell.coordinate].fill.fgColor.rgb != '00000000':
                print(sheet[cell.coordinate].fill.fgColor.rgb)

            # TODO 色修正はまたこんど。ログででてきたやつ
            # 00000000
            # FFD6FEE4
            # FFFFE7FF
            # FFFFFF00


# 結合セルを解除簡略版。休日セル想定
def unmergeRestCell(sheet, start_row, start_column):

    unmergeCell(sheet, start_row, start_column, start_row, start_column + 4)

# 結合セルを解除
def unmergeCell(sheet, start_row, start_column, end_row, end_column):

    sheet.unmerge_cells(start_row = start_row, start_column = start_column, end_row = end_row, end_column = end_column)


if __name__ == '__main__':
    main()
